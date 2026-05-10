"""Fyll ut kundens Excel-prisskjema (Bilag 7)."""
from openpyxl import load_workbook
from pathlib import Path

SRC = Path("/Users/asbis/code/asbisdev/oppdrag/aktive/2026-105336-rusinfo-selvhjelpsapp/fra-kunden/02-ssa-t-kontrakt/03.2 DEL 2 SSA-T PRISSKJEMA Vedlegg 1 til bilag 7.xlsx")
DST = Path("/Users/asbis/code/asbisdev/oppdrag/aktive/2026-105336-rusinfo-selvhjelpsapp/til-innsending/autogenerert/03.2 DEL 2 SSA-T PRISSKJEMA Vedlegg 1 til bilag 7 (utfylt).xlsx")

wb = load_workbook(SRC)
ws = wb["Pris for tjenesten"]

# 1.1 Hovedoppdrag
ws["F11"] = (
    "Utvikling av selvhjelpsapp for iOS og Android i React Native + TypeScript. "
    "Inkluderer design, utvikling, testing (unit, e2e, manuell), intern og ekstern "
    "UU-revisjon (Funka Nu eller MediaLT), Strapi-backend, CI/CD, publisering i "
    "App Store og Google Play under Kundens kontoer, sluttbruker- og admin-manualer, "
    "samt personvernvurdering. Kildekoden overføres til Kunden fra dag 1 via felles "
    "git-repo. Fastpris, inkluderer alt frem til leveringsdag."
)
ws["G11"] = 650000

# 2.1 Vedlikehold
ws["F15"] = (
    "Løpende drift og vedlikehold: feilretting, sikkerhetsoppdateringer, plattform-"
    "oppgraderinger (årlig iOS/Android SDK), Strapi-patching, overvåkning, "
    "backup-kontroll, feilmottak via app og e-post. Responstid 24 timer på virkedager. "
    "Inkluderer ikke videreutvikling av nye funksjoner (prises etter timepris 6.1/6.2)."
)
ws["G15"] = 120000

# 6.1 Junior timepris
ws["F19"] = "Leverandørens hovedressurs er senior (jf. rad 6.2). Skulle Kunden ha behov for en juniorressurs til oppgaver som ikke krever seniorkompetanse, hentes denne inn fra Leverandørens nettverk (3–5 års React Native-erfaring) og pre-godkjennes skriftlig av Kunden før oppstart."
ws["G19"] = 1100

# 6.2 Senior timepris
ws["F20"] = "Asbjørn Rørvik — senior fullstack-utvikler med 8+ års produksjonserfaring. Hovedressurs og arkitekt på prosjektet."
ws["G20"] = 1400

# 4. Opplæring (rad 24)
ws["F24"] = (
    "Oppstarts-workshop (4 timer, fysisk i Oslo eller Teams) med gjennomgang av "
    "Strapi-administrasjon: innholdsredigering, dagens tema, default kriseplan, "
    "statistikkhenting. Inkluderer skriftlig admin-manual (PDF + Markdown i git, "
    "med skjermbilder) og 2 timers oppfølgingssamtale ved behov innen 30 dager "
    "etter leveringsdag."
)
ws["G24"] = 15000

# Øvrig sortiment — hopp over, fyll manuelt ved behov. Ikke del av evalueringspris.

DST.parent.mkdir(parents=True, exist_ok=True)
wb.save(DST)

# Print oppsummering
print("Fyllt ut:")
print(f"  Rad 11 (1.1 Hovedoppdrag):  650 000 NOK")
print(f"  Rad 15 (2.1 Vedlikehold):   120 000 NOK/år × 10 år = 1 200 000 NOK")
print(f"  Rad 19 (6.1 Junior):        1 100 NOK/t × 20 t =  22 000 NOK")
print(f"  Rad 20 (6.2 Senior):        1 400 NOK/t × 10 t =  14 000 NOK")
print(f"  Rad 24 (Opplæring):          15 000 NOK")
print(f"  ---")
print(f"  Samlet evalueringspris:   1 901 000 NOK")
print(f"\nLagret til: {DST}")
